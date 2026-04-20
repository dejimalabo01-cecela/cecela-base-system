import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from datetime import date
import calendar

# ── 設定 ──────────────────────────────────────────────────────
TIMELINE_START = (2025, 10)   # タイムライン開始年月
NUM_MONTHS = 24               # 表示月数

TASKS = [
    ('PJ設定期日',                '10B981'),
    ('仕入れ決済日',              '34D399'),
    ('解体工事',                  'F87171'),
    ('造成工事',                  'FCD34D'),
    ('上下水・ガス引込・道路復旧',  '60A5FA'),
    ('プラン作成',                'A78BFA'),
    ('建築確認',                  '22D3EE'),
    ('建築工事（外構含む）',       'FB923C'),
    ('販売',                      'F472B6'),
]

ROW_PROP     = 1   # 物件情報
ROW_HEADER   = 2   # 月ヘッダー
ROW_WEEK     = 3   # 週ラベル
ROW_D_START  = 4   # スロット開始日（非表示）
ROW_D_END    = 5   # スロット終了日（非表示）
ROW_TASK_1   = 6   # タスク1行目

COL_TASK  = 1   # A: 工程名
COL_START = 2   # B: 開始日
COL_END   = 3   # C: 終了日
COL_GANTT = 4   # D: ガント開始列

# ── スロット生成 ───────────────────────────────────────────────
def gen_slots(sy, sm, n):
    res = []
    y, m = sy, sm
    for _ in range(n):
        last = calendar.monthrange(y, m)[1]
        for ws, we in [(1,7),(8,14),(15,21),(22,last)]:
            res.append((date(y,m,ws), date(y,m,we), y, m))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return res

slots = gen_slots(*TIMELINE_START, NUM_MONTHS)
END_COL  = COL_GANTT + len(slots) - 1
S_COL    = get_column_letter(COL_GANTT)
E_COL    = get_column_letter(END_COL)

# ── ワークブック作成 ───────────────────────────────────────────
wb = openpyxl.Workbook()

def make_side(style='thin', color='E2E8F0'):
    return Side(style=style, color=color)

def bdr(left=None, right=None, top=None, bottom=None):
    return Border(left=left, right=right, top=top, bottom=bottom)

# ══════════════════════════════════════════════════════════════
# 物件一覧 シート
# ══════════════════════════════════════════════════════════════
wl = wb.active
wl.title = '物件一覧'
wl.sheet_view.showGridLines = True

BLUE_FILL  = PatternFill(start_color='1E40AF', fill_type='solid')
WHT_BOLD   = Font(color='FFFFFF', bold=True, name='Meiryo UI', size=10)
T          = make_side()
FULL_BDR   = Border(left=T, right=T, top=T, bottom=T)

for i, h in enumerate(['物件ID', '物件名', '登録日'], 1):
    c = wl.cell(1, i, h)
    c.fill = BLUE_FILL;  c.font = WHT_BOLD
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = FULL_BDR

wl.column_dimensions['A'].width = 10
wl.column_dimensions['B'].width = 35
wl.column_dimensions['C'].width = 14
wl.row_dimensions[1].height = 22

note = wl.cell(2, 1, '↑ 新規物件を追加したらここに記入してください')
note.font = Font(color='94A3B8', name='Meiryo UI', size=9, italic=True)

# ══════════════════════════════════════════════════════════════
# [テンプレート] シート
# ══════════════════════════════════════════════════════════════
ws = wb.create_sheet('★テンプレート★')
ws.sheet_view.showGridLines = False

DARK   = PatternFill(start_color='1E3A5F', fill_type='solid')
MID_D  = PatternFill(start_color='2D4E7A', fill_type='solid')
GRAY   = PatternFill(start_color='F1F5F9', fill_type='solid')
WHITE  = PatternFill(start_color='FFFFFF', fill_type='solid')
ALT    = PatternFill(start_color='F8FAFC', fill_type='solid')

# ── 行1: 物件情報 ─────────────────────────────────────────────
ws.row_dimensions[ROW_PROP].height = 28
ws.cell(ROW_PROP, COL_TASK, '物件ID').font = Font(bold=True, name='Meiryo UI', size=10)
ws.cell(ROW_PROP, COL_START, 'P-XXX').font = Font(bold=True, name='Meiryo UI', size=12, color='1D4ED8')
ws.cell(ROW_PROP, COL_END, '物件名:').font = Font(bold=True, name='Meiryo UI', size=10)
ws.cell(ROW_PROP, COL_GANTT, '（ここに物件名を入力）').font = Font(name='Meiryo UI', size=11, color='94A3B8')

for col in [COL_TASK, COL_START, COL_END, COL_GANTT]:
    ws.cell(ROW_PROP, col).alignment = Alignment(vertical='center')

# ── 行2: 月ヘッダー ────────────────────────────────────────────
ws.row_dimensions[ROW_HEADER].height = 18
for col, text in [(COL_TASK,'工程'),(COL_START,'開始日'),(COL_END,'終了日')]:
    c = ws.cell(ROW_HEADER, col, text)
    c.fill = DARK
    c.font = Font(color='FFFFFF', bold=True, name='Meiryo UI', size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')

# 月グループ収集
month_groups = []
prev_key = None; msc = None; mlabel = None
for i, (_, _, y, m) in enumerate(slots):
    col = COL_GANTT + i
    key = (y, m)
    if key != prev_key:
        if prev_key is not None:
            month_groups.append((msc, col-1, mlabel))
        msc = col; prev_key = key
        if i == 0 or slots[i-1][2] != y:
            mlabel = f'{y}/{m}月'
        else:
            mlabel = f'{m}月'
if prev_key:
    month_groups.append((msc, END_COL, mlabel))

for sc, ec, label in month_groups:
    c = ws.cell(ROW_HEADER, sc, label)
    c.fill = DARK
    c.font = Font(color='FFFFFF', bold=True, name='Meiryo UI', size=8)
    c.alignment = Alignment(horizontal='center', vertical='center')
    if sc < ec:
        ws.merge_cells(start_row=ROW_HEADER, start_column=sc, end_row=ROW_HEADER, end_column=ec)

# ── 行3: 週ラベル ─────────────────────────────────────────────
ws.row_dimensions[ROW_WEEK].height = 11
for col in [COL_TASK, COL_START, COL_END]:
    ws.cell(ROW_WEEK, col).fill = DARK

for i, (sd, _, y, m) in enumerate(slots):
    col = COL_GANTT + i
    wn = (sd.day - 1) // 7 + 1
    c = ws.cell(ROW_WEEK, col, f'{wn}週')
    c.fill = MID_D
    c.font = Font(color='CBD5E1', name='Meiryo UI', size=6)
    c.alignment = Alignment(horizontal='center', vertical='center')

# ── 行4-5: スロット日付（非表示） ─────────────────────────────
for i, (ss, se, y, m) in enumerate(slots):
    col = COL_GANTT + i
    ws.cell(ROW_D_START, col, ss).number_format = 'YYYY/MM/DD'
    ws.cell(ROW_D_END, col, se).number_format = 'YYYY/MM/DD'

ws.row_dimensions[ROW_D_START].hidden = True
ws.row_dimensions[ROW_D_END].hidden = True

# ── 条件付き書式: Today ハイライト（低優先度）─────────────────
ALL_GANTT = f'{S_COL}{ROW_TASK_1}:{E_COL}{ROW_TASK_1 + len(TASKS) - 1}'
today_formula = [f'AND({S_COL}${ROW_D_START}<=TODAY(),TODAY()<={S_COL}${ROW_D_END})']
today_fill = PatternFill(start_color='DBEAFE', fill_type='solid')
ws.conditional_formatting.add(ALL_GANTT, FormulaRule(formula=today_formula, fill=today_fill))

# ── 条件付き書式: タスクバー（高優先度）─────────────────────────
for task_idx, (task_name, color_hex) in enumerate(TASKS):
    row = ROW_TASK_1 + task_idx
    task_range = f'{S_COL}{row}:{E_COL}{row}'
    formula = [
        f'AND($B{row}>=1,$C{row}>=1,'
        f'$B{row}<={S_COL}${ROW_D_END},'
        f'$C{row}>={S_COL}${ROW_D_START})'
    ]
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
    ws.conditional_formatting.add(task_range, FormulaRule(formula=formula, fill=fill))

# ── タスク行スタイル ───────────────────────────────────────────
for task_idx, (task_name, color_hex) in enumerate(TASKS):
    row = ROW_TASK_1 + task_idx
    ws.row_dimensions[row].height = 22

    # 工程名セル
    c = ws.cell(row, COL_TASK, task_name)
    c.font = Font(name='Meiryo UI', size=9, bold=True, color='1E293B')
    c.fill = GRAY
    c.alignment = Alignment(vertical='center', indent=1)
    c.border = Border(
        left=Side(style='medium', color=color_hex),
        right=make_side(), top=make_side(), bottom=make_side()
    )

    # 日付入力セル
    for col in [COL_START, COL_END]:
        c = ws.cell(row, col)
        c.number_format = 'YYYY/MM/DD'
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = WHITE
        c.border = Border(left=make_side(), right=make_side(), top=make_side(), bottom=make_side())

    # ガントセル
    for i in range(len(slots)):
        col = COL_GANTT + i
        _, _, _, m = slots[i]
        c = ws.cell(row, col)
        c.fill = ALT if m % 2 == 0 else WHITE
        is_month_end = (i + 1 < len(slots) and slots[i+1][3] != m) or i == len(slots) - 1
        right_s = make_side('thin', 'CBD5E1') if is_month_end else make_side('dotted', 'E5E7EB')
        c.border = Border(right=right_s, top=make_side('thin','F1F5F9'), bottom=make_side('thin','F1F5F9'))

# ── 列幅 ─────────────────────────────────────────────────────
ws.column_dimensions[get_column_letter(COL_TASK)].width  = 26
ws.column_dimensions[get_column_letter(COL_START)].width = 13
ws.column_dimensions[get_column_letter(COL_END)].width   = 13
for i in range(len(slots)):
    ws.column_dimensions[get_column_letter(COL_GANTT + i)].width = 3.8

# ── ペインの固定 ──────────────────────────────────────────────
ws.freeze_panes = ws.cell(ROW_TASK_1, COL_GANTT)

# ══════════════════════════════════════════════════════════════
# 使い方 シート
# ══════════════════════════════════════════════════════════════
wi = wb.create_sheet('使い方')
wi.column_dimensions['A'].width = 65
wi.sheet_view.showGridLines = False

lines = [
    ('Cecela 物件進捗ガントチャート  使い方', True, 13),
    ('', False, 10),
    ('■ 新規物件を追加するには', True, 10),
    ('  1. [テンプレート] シートタブを 右クリック → 「移動またはコピー」', False, 10),
    ('  2. 「コピーを作成する」 にチェックを入れて OK', False, 10),
    ('  3. コピーされたタブをダブルクリックして名前変更（例: P-001）', False, 10),
    ('  4. B1セルに 物件ID（例: P-001）を入力', False, 10),
    ('  5. D1セルに 物件名を入力', False, 10),
    ('', False, 10),
    ('■ スケジュールを入力するには', True, 10),
    ('  各工程の「開始日」「終了日」セルに日付を入力', False, 10),
    ('  → 入力するだけで、ガントチャートに自動で色がつきます', False, 10),
    ('  ※ 入力形式: 2026/05/10 または 2026-05-10', False, 10),
    ('', False, 10),
    ('■ 工程の色について', True, 10),
]

for i, (text, bold, size) in enumerate(lines, 1):
    c = wi.cell(i, 1, text)
    c.font = Font(name='Meiryo UI', size=size, bold=bold, color='1E293B' if bold else '374151')
    wi.row_dimensions[i].height = 20

offset = len(lines) + 2
for j, (task_name, color_hex) in enumerate(TASKS):
    row = offset + j
    c = wi.cell(row, 1, f'   {task_name}')
    c.fill = PatternFill(start_color=color_hex, fill_type='solid')
    c.font = Font(name='Meiryo UI', size=10, color='FFFFFF', bold=True)
    c.alignment = Alignment(vertical='center')
    wi.row_dimensions[row].height = 22

# ══════════════════════════════════════════════════════════════
# 保存
# ══════════════════════════════════════════════════════════════
out = r'C:\Users\endle\OneDrive\デスクトップ\cecela-base-system-pj\cecela-gantt.xlsx'
wb.save(out)
print(f'OK: {out}')
print(f'Months: {NUM_MONTHS}, Slots: {len(slots)}, Last col: {E_COL}')
